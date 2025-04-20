from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import xlsxwriter

class ReportGenerator:
    def __init__(self):
        self.header_text = [
            "UNIVERSITY OF CEBU MAIN CAMPUS",
            "COLLEGE OF COMPUTER STUDIES",
            "Laboratory Management System"
        ]

    def _add_header(self, canvas, width):
        # Set up header style
        canvas.setFont("Helvetica-Bold", 14)
        y_position = 800
        
        # Add header text centered
        for text in self.header_text:
            canvas.drawCentredString(width/2, y_position, text)
            y_position -= 20
        
        # Add horizontal line
        canvas.line(50, y_position-10, width-50, y_position-10)
        return y_position-30  # Return the new y position after header

    def _format_filters(self, filters):
        if not filters:
            return []
        formatted = []
        for key, value in filters.items():
            if value:
                formatted.append(f"{key}: {value}")
        return formatted

    def _prepare_table_data(self, data):
        # Handle both dictionary and list formats
        if isinstance(data, dict):
            headers = data.get('headers', [])
            rows = data.get('rows', [])
        elif isinstance(data, list):
            # If data is a list, first row is headers, rest are data
            if data:
                headers = data[0]
                rows = data[1:]
            else:
                headers = []
                rows = []
        else:
            headers = []
            rows = []
        return headers, rows

    def generate_pdf(self, title, data, filters=None):
        buffer = BytesIO()
        
        # Create the PDF object
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Add header
        styles = getSampleStyleSheet()
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add university header
        for header in self.header_text:
            elements.append(Paragraph(header, header_style))
        
        # Add title
        elements.append(Paragraph(title, header_style))
        elements.append(Spacer(1, 20))
        
        # Add filters if provided
        if filters:
            filter_style = ParagraphStyle(
                'Filters',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=20
            )
            filter_text = " | ".join(self._format_filters(filters))
            elements.append(Paragraph(f"Filters: {filter_text}", filter_style))
            elements.append(Spacer(1, 20))
        
        # Create table for data
        headers, rows = self._prepare_table_data(data)
        if headers and rows:
            # Calculate column widths based on content
            col_widths = self._calculate_column_widths(headers, rows)
            
            table_data = [headers] + rows
            table = Table(table_data, colWidths=col_widths)
            
            # Define table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ])
            table.setStyle(table_style)
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _calculate_column_widths(self, headers, rows):
        # Get available width (letter page width - margins)
        available_width = letter[0] - 144  # 72 points margin on each side
        
        # Calculate maximum width needed for each column
        max_widths = []
        for i in range(len(headers)):
            col_data = [str(row[i]) for row in rows]
            col_data.append(str(headers[i]))
            max_width = max(len(data) for data in col_data) * 7  # Approximate width per character
            max_widths.append(max_width)
        
        # Adjust widths to fit page
        total_width = sum(max_widths)
        if total_width > available_width:
            # Scale down proportionally
            scale = available_width / total_width
            max_widths = [width * scale for width in max_widths]
        
        return max_widths

    def generate_excel(self, title, data, filters=None):
        buffer = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        
        # Add header
        current_row = 1
        for header in self.header_text:
            worksheet.merge_cells(f'A{current_row}:G{current_row}')
            cell = worksheet.cell(row=current_row, column=1, value=header)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Add title
        current_row += 1
        worksheet.merge_cells(f'A{current_row}:G{current_row}')
        cell = worksheet.cell(row=current_row, column=1, value=title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center')
        
        # Add filters if provided
        if filters:
            current_row += 1
            filter_text = " | ".join(self._format_filters(filters))
            worksheet.merge_cells(f'A{current_row}:G{current_row}')
            cell = worksheet.cell(row=current_row, column=1, value=f"Filters: {filter_text}")
            cell.font = Font(italic=True)
        
        # Add data
        headers, rows = self._prepare_table_data(data)
        if headers and rows:
            current_row += 2
            # Headers
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_data in rows:
                current_row += 1
                for col, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_csv(self, title, data, filters=None):
        buffer = StringIO()
        writer = csv.writer(buffer)
        
        # Write header
        for header in self.header_text:
            writer.writerow([header])
        
        writer.writerow([])  # Empty row
        writer.writerow([title])
        
        # Write filters if provided
        if filters:
            filter_text = " | ".join(self._format_filters(filters))
            writer.writerow([f"Filters: {filter_text}"])
            writer.writerow([])  # Empty row
        
        # Write data
        headers, rows = self._prepare_table_data(data)
        if headers and rows:
            writer.writerow(headers)
            writer.writerows(rows)
        
        # Get the string value and encode it with BOM for Excel compatibility
        csv_data = buffer.getvalue()
        # Add BOM for Excel compatibility
        csv_data_with_bom = '\ufeff' + csv_data
        
        # Create a new BytesIO object with the encoded data
        bytes_buffer = BytesIO()
        bytes_buffer.write(csv_data_with_bom.encode('utf-8-sig'))
        bytes_buffer.seek(0)
        return bytes_buffer 