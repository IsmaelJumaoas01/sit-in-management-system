from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for, send_file, Response
from db import get_db_connection
import os
from werkzeug.utils import secure_filename
import base64
import pandas as pd
from io import BytesIO, StringIO
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import xlsxwriter
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils.report_generator import ReportGenerator

staff_bp = Blueprint('staff', __name__)

# Add configuration for file uploads
UPLOAD_FOLDER = 'static/profile_pictures'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@staff_bp.route('/')
@staff_bp.route('/dashboard')
def staff_dashboard():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return redirect(url_for('auth.login'))

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get all laboratories
        cursor.execute("SELECT * FROM LABORATORIES")
        laboratories = cursor.fetchall()

        # Get staff details
        cursor.execute("""
            SELECT FIRSTNAME, LASTNAME, EMAIL 
            FROM USERS 
            WHERE IDNO = %s
        """, (session['IDNO'],))
        staff_data = cursor.fetchone()

        return render_template('staff_dashboard.html', 
                             laboratories=[{
                                 'LAB_ID': lab[0],
                                 'LAB_NAME': lab[1]
                             } for lab in laboratories],
                             staff_firstname=staff_data[0],
                             staff_lastname=staff_data[1],
                             staff_email=staff_data[2])
    except Exception as e:
        flash(str(e), 'error')
        return redirect(url_for('auth.login'))
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/edit_info', methods=['GET', 'POST'])
def edit_info():
    if 'IDNO' in session and session['USER_TYPE'] == 'STAFF':
        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            # Get all laboratories (needed for the template)
            cursor.execute("SELECT * FROM LABORATORIES")
            laboratories = cursor.fetchall()

            # Get current staff data
            cursor.execute("""
                SELECT FIRSTNAME, LASTNAME, EMAIL 
                FROM USERS 
                WHERE IDNO = %s
            """, (session['IDNO'],))
            staff_data = cursor.fetchone()

            if request.method == 'GET':
                return render_template('staff_dashboard.html',
                                    laboratories=[{
                                        'LAB_ID': lab[0],
                                        'LAB_NAME': lab[1]
                                    } for lab in laboratories],
                                    staff_firstname=staff_data[0],
                                    staff_lastname=staff_data[1],
                                    staff_email=staff_data[2])
            
            elif request.method == 'POST':
                new_firstname = request.form['firstname']
                new_lastname = request.form['lastname']
                new_email = request.form['email']
                new_password = request.form.get('password')

                try:
                    if new_password:
                        cursor.execute("""
                            UPDATE USERS 
                            SET FIRSTNAME = %s, LASTNAME = %s, EMAIL = %s, PASSWORD = %s 
                            WHERE IDNO = %s
                        """, (new_firstname, new_lastname, new_email, new_password, session['IDNO']))
                    else:
                        cursor.execute("""
                            UPDATE USERS 
                            SET FIRSTNAME = %s, LASTNAME = %s, EMAIL = %s 
                            WHERE IDNO = %s
                        """, (new_firstname, new_lastname, new_email, session['IDNO']))

                    conn.commit()
                    
                    # Update session data
                    session['FIRSTNAME'] = new_firstname
                    session['LASTNAME'] = new_lastname
                    session['EMAIL'] = new_email

                    flash('Your details have been updated successfully!', 'success')
                except Exception as e:
                    flash(f'An error occurred: {str(e)}', 'error')

                return redirect(url_for('staff.staff_dashboard'))

        except Exception as e:
            flash(str(e), 'error')
            return redirect(url_for('auth.login'))
        finally:
            cursor.close()
            conn.close()

    flash('You must be logged in as staff to edit information.', 'error')
    return redirect(url_for('auth.login'))

@staff_bp.route('/update_profile_picture', methods=['POST'])
def update_profile_picture():
    if 'IDNO' not in session or session.get('USER_TYPE') != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    if 'profile_picture' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['profile_picture']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if file and allowed_file(file.filename):
        # Read the file content
        file_content = file.read()
        
        # Update the user's profile picture in the database
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "UPDATE USERS SET PROFILE_PICTURE = %s WHERE IDNO = %s",
                (file_content, session['IDNO'])
            )
            conn.commit()
            
            # Convert image to base64 for immediate display
            encoded_image = base64.b64encode(file_content).decode('utf-8')
            
            return jsonify({
                'success': True,
                'profile_picture_data': f"data:image/{file.filename.split('.')[-1]};base64,{encoded_image}"
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            conn.close()

    return jsonify({'error': 'Invalid file type'}), 400

@staff_bp.route('/start_sitin', methods=['POST'])
def start_sitin():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    student_id = data.get('student_id')
    lab_id = data.get('lab_id')
    purpose_id = data.get('purpose_id')

    if not all([student_id, lab_id, purpose_id]):
        return jsonify({'error': 'Missing required fields'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Check if student has an ongoing sit-in session
        cursor.execute("""
            SELECT COUNT(*) FROM SIT_IN_RECORDS 
            WHERE USER_IDNO = %s AND SESSION = 'ON_GOING'
        """, (student_id,))
        ongoing_sessions = cursor.fetchone()[0]
        if ongoing_sessions > 0:
            return jsonify({'error': 'Student already has an ongoing sit-in session'}), 400

        # Check student's sit-in limit
        cursor.execute("SELECT SIT_IN_COUNT FROM SIT_IN_LIMITS WHERE USER_IDNO = %s", (student_id,))
        result = cursor.fetchone()
        if not result or result[0] <= 0:
            return jsonify({'error': 'Student has no remaining sit-in sessions'}), 400

        # Get an available computer in the lab
        cursor.execute("""
            SELECT c.COMPUTER_ID 
            FROM COMPUTERS c
            LEFT JOIN SIT_IN_RECORDS s ON c.COMPUTER_ID = s.COMPUTER_ID AND s.SESSION = 'ON_GOING'
            WHERE c.LAB_ID = %s AND s.COMPUTER_ID IS NULL
            LIMIT 1
        """, (lab_id,))
        computer = cursor.fetchone()
        if not computer:
            return jsonify({'error': 'No available computers in this lab'}), 400

        # Start the sit-in session
        cursor.execute("""
            INSERT INTO SIT_IN_RECORDS 
            (USER_IDNO, LAB_ID, COMPUTER_ID, PURPOSE_ID, DATE, STATUS, SESSION)
            VALUES (%s, %s, %s, %s, NOW(), 'APPROVED', 'ON_GOING')
        """, (student_id, lab_id, computer[0], purpose_id))

        conn.commit()
        return jsonify({'success': True, 'message': 'Sit-in session started successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/active_sessions')
def get_active_sessions():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    lab_id = request.args.get('lab_id')
    purpose_id = request.args.get('purpose_id')
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT s.RECORD_ID, s.USER_IDNO, s.LAB_ID, s.DATE, s.END_TIME,
                   u.FIRSTNAME, u.LASTNAME,
                   l.LAB_NAME,
                   p.PURPOSE_NAME
            FROM SIT_IN_RECORDS s
            JOIN USERS u ON s.USER_IDNO = u.IDNO
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            WHERE s.SESSION = 'ON_GOING'
        """
        params = []

        if lab_id:
            query += " AND s.LAB_ID = %s"
            params.append(lab_id)
        
        if purpose_id:
            query += " AND s.PURPOSE_ID = %s"
            params.append(purpose_id)

        print("Executing active sessions query:", query)
        print("With parameters:", params)
        
        cursor.execute(query, params)
        sessions = cursor.fetchall()
        print(f"Found {len(sessions)} active sessions")

        result = []
        for s in sessions:
            profile_url = url_for('user.get_profile_picture', idno=s[1], _external=True)
            print(f"Generated profile URL for student {s[1]}: {profile_url}")
            
            session_data = {
                'RECORD_ID': s[0],
                'STUDENT_ID': s[1],
                'LAB_ID': s[2],
                'DATE': s[3].isoformat() if s[3] else None,
                'END_TIME': s[4].isoformat() if s[4] else None,
                'STUDENT_NAME': f"{s[5]} {s[6]}",
                'LAB_NAME': s[7],
                'PURPOSE_NAME': s[8],
                'PROFILE_PICTURE_URL': profile_url
            }
            print(f"Processing session for student {s[1]}: {session_data}")
            result.append(session_data)
        
        print("Returning active sessions data:", result)
        return jsonify(result)
    except Exception as e:
        print(f"Error in active_sessions: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/today_ended_sessions')
def today_ended_sessions():
    print("\n!!!! ROUTE /today_ended_sessions WAS CALLED !!!!")  # Initial debug message
    
    if 'IDNO' not in session or session.get('USER_TYPE') != 'STAFF':
        print("Unauthorized access attempt")
        return jsonify({'error': 'Unauthorized'}), 401

    print("Staff authorized, proceeding with database query")
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        today = datetime.now().strftime('%Y-%m-%d')
        print(f"Checking ended sessions for date: {today}")

        # Simple query to check database connection
        test_query = "SELECT COUNT(*) FROM SIT_IN_RECORDS"
        cursor.execute(test_query)
        total_records = cursor.fetchone()[0]
        print(f"Total records in SIT_IN_RECORDS: {total_records}")

        # Main query
        query = """
            SELECT 
                s.RECORD_ID,
                s.USER_IDNO,
                CONCAT(u.FIRSTNAME, ' ', u.LASTNAME) as STUDENT_NAME,
                l.LAB_NAME,
                s.DATE,
                s.END_TIME,
                p.PURPOSE_NAME,
                s.SESSION
            FROM SIT_IN_RECORDS s
            JOIN USERS u ON s.USER_IDNO = u.IDNO
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            WHERE DATE(s.END_TIME) = %s
            AND s.SESSION = 'ENDED'
            ORDER BY s.END_TIME DESC
        """
        
        print("About to execute main query")
        cursor.execute(query, (today,))
        rows = cursor.fetchall()
        print(f"Query returned {len(rows)} rows")

        ended_sessions = []
        for row in rows:
            try:
                profile_url = url_for('user.get_profile_picture', idno=row[1], _external=True)
                session_data = {
                    'record_id': row[0],
                    'user_idno': row[1],
                    'student_name': row[2],
                    'lab_name': row[3],
                    'date': row[4].strftime("%Y-%m-%d %H:%M:%S") if row[4] else None,
                    'end_time': row[5].strftime("%Y-%m-%d %H:%M:%S") if row[5] else None,
                    'purpose_name': row[6],
                    'status': row[7],
                    'PROFILE_PICTURE_URL': profile_url
                }
                ended_sessions.append(session_data)
                print(f"Processed session ID: {row[0]}")
            except Exception as e:
                print(f"Error processing row: {str(e)}")

        print(f"Returning {len(ended_sessions)} sessions")
        return jsonify(ended_sessions)

    except Exception as e:
        print(f"ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()
        print("Database connection closed")

@staff_bp.route('/end_session/<int:record_id>', methods=['POST'])
def end_session(record_id):
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get the student ID for the session
        cursor.execute("SELECT USER_IDNO FROM SIT_IN_RECORDS WHERE RECORD_ID = %s", (record_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({'error': 'Session not found'}), 404

        student_id = result[0]

        # Update the session status and set END_TIME
        cursor.execute("""
            UPDATE SIT_IN_RECORDS 
            SET SESSION = 'ENDED', END_TIME = NOW()
            WHERE RECORD_ID = %s
        """, (record_id,))

        # Decrease the student's sit-in limit
        cursor.execute("""
            UPDATE SIT_IN_LIMITS 
            SET SIT_IN_COUNT = SIT_IN_COUNT - 1 
            WHERE USER_IDNO = %s
        """, (student_id,))

        conn.commit()
        return jsonify({'success': True, 'message': 'Session ended successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/laboratories')
def get_laboratories():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT LAB_ID, LAB_NAME FROM LABORATORIES")
        labs = cursor.fetchall()
        return jsonify([{
            'LAB_ID': lab[0],
            'LAB_NAME': lab[1]
        } for lab in labs])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/purposes')
def get_purposes():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT PURPOSE_ID, PURPOSE_NAME FROM PURPOSES ORDER BY PURPOSE_NAME")
        purposes = cursor.fetchall()
        return jsonify([{
            'PURPOSE_ID': purpose[0],
            'PURPOSE_NAME': purpose[1]
        } for purpose in purposes])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/purposes/<int:purpose_id>', methods=['PUT'])
def update_purpose(purpose_id):
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json()
    purpose_name = data.get('purposeName')

    if not purpose_name:
        return jsonify({'error': 'Purpose name is required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Check if purpose exists
        cursor.execute("SELECT PURPOSE_ID FROM PURPOSES WHERE PURPOSE_ID = %s", (purpose_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Purpose not found'}), 404

        # Check if new name already exists (excluding current purpose)
        cursor.execute("SELECT PURPOSE_ID FROM PURPOSES WHERE PURPOSE_NAME = %s AND PURPOSE_ID != %s", 
                      (purpose_name, purpose_id))
        if cursor.fetchone():
            return jsonify({'error': 'A purpose with this name already exists'}), 400

        # Update purpose
        cursor.execute("UPDATE PURPOSES SET PURPOSE_NAME = %s WHERE PURPOSE_ID = %s",
                      (purpose_name, purpose_id))
        conn.commit()
        
        return jsonify({'success': True, 'message': 'Purpose updated successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/check_remaining_sessions/<string:student_id>')
def check_remaining_sessions(student_id):
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # First check if student exists and is actually a student
        cursor.execute("""
            SELECT IDNO FROM USERS 
            WHERE IDNO = %s AND USER_TYPE = 'STUDENT'
        """, (student_id,))
        
        if not cursor.fetchone():
            return jsonify({
                'error': 'Student not found',
                'remaining_sessions': 0
            }), 404

        # Get remaining sessions
        cursor.execute("""
            SELECT SIT_IN_COUNT 
            FROM SIT_IN_LIMITS 
            WHERE USER_IDNO = %s
        """, (student_id,))
        
        result = cursor.fetchone()
        if not result:
            return jsonify({
                'error': 'No sit-in limit found for student',
                'remaining_sessions': 0
            }), 404

        return jsonify({
            'remaining_sessions': result[0]
        })

    except Exception as e:
        return jsonify({
            'error': str(e),
            'remaining_sessions': 0
        }), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/sitin_records')
def get_sitin_records():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    lab_id = request.args.get('lab_id')
    purpose_id = request.args.get('purpose_id')
    status = request.args.get('status')
    session_status = request.args.get('session')
    student_id = request.args.get('student_id')
    
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT s.RECORD_ID, s.USER_IDNO, s.LAB_ID, s.DATE, s.END_TIME, s.STATUS, s.SESSION,
                   u.FIRSTNAME, u.LASTNAME,
                   l.LAB_NAME,
                   p.PURPOSE_NAME
            FROM SIT_IN_RECORDS s
            JOIN USERS u ON s.USER_IDNO = u.IDNO
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            WHERE 1=1
        """
        params = []

        if lab_id:
            query += " AND s.LAB_ID = %s"
            params.append(lab_id)
        
        if purpose_id:
            query += " AND s.PURPOSE_ID = %s"
            params.append(purpose_id)
            
        if status:
            query += " AND s.STATUS = %s"
            params.append(status)
            
        if session_status:
            query += " AND s.SESSION = %s"
            params.append(session_status)

        if student_id:
            query += " AND s.USER_IDNO = %s"
            params.append(student_id)

        query += " ORDER BY s.DATE DESC"
        
        cursor.execute(query, params)
        records = cursor.fetchall()

        return jsonify([{
            'RECORD_ID': r[0],
            'STUDENT_ID': r[1],
            'LAB_ID': r[2],
            'DATE': r[3].isoformat(),
            'END_TIME': r[4].isoformat() if r[4] else None,
            'STATUS': r[5],
            'SESSION': r[6],
            'STUDENT_NAME': f"{r[7]} {r[8]}",
            'LAB_NAME': r[9],
            'PURPOSE_NAME': r[10]
        } for r in records])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/statistics')
def get_statistics():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        print("Unauthorized access attempt to staff statistics")
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        print("Starting staff statistics retrieval...")
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get registered students count
        cursor.execute("SELECT COUNT(IDNO) FROM USERS WHERE USER_TYPE = 'STUDENT'")
        registered_students = cursor.fetchone()[0] or 0
        print(f"Found {registered_students} registered students")

        # Get current sit-ins count
        cursor.execute("""
            SELECT COUNT(RECORD_ID) 
            FROM SIT_IN_RECORDS 
            WHERE SESSION = 'ON_GOING'
        """)
        current_sitins = cursor.fetchone()[0] or 0
        print(f"Found {current_sitins} current sit-ins")

        # Get total sit-ins count
        cursor.execute("SELECT COUNT(RECORD_ID) FROM SIT_IN_RECORDS")
        total_sitins = cursor.fetchone()[0] or 0
        print(f"Found {total_sitins} total sit-ins")

        # Get sit-ins by purpose with proper join
        cursor.execute("""
            SELECT 
                p.PURPOSE_NAME,
                COALESCE(COUNT(s.RECORD_ID), 0) as count
            FROM PURPOSES p
            LEFT JOIN SIT_IN_RECORDS s ON p.PURPOSE_ID = s.PURPOSE_ID
            GROUP BY p.PURPOSE_ID, p.PURPOSE_NAME
            ORDER BY count DESC
        """)
        purpose_stats = []
        for row in cursor.fetchall():
            if row[0]:  # Only add if purpose name exists
                purpose_stats.append({
                    'purpose_name': row[0],
                    'count': int(row[1])
                })
        print(f"Found purpose stats: {purpose_stats}")

        # Get sit-ins by lab with proper join
        cursor.execute("""
            SELECT 
                l.LAB_NAME,
                COALESCE(COUNT(s.RECORD_ID), 0) as count
            FROM LABORATORIES l
            LEFT JOIN SIT_IN_RECORDS s ON l.LAB_ID = s.LAB_ID
            GROUP BY l.LAB_ID, l.LAB_NAME
            ORDER BY count DESC
        """)
        lab_stats = []
        for row in cursor.fetchall():
            if row[0]:  # Only add if lab name exists
                lab_stats.append({
                    'lab_name': row[0],
                    'count': int(row[1])
                })
        print(f"Found lab stats: {lab_stats}")

        cursor.close()
        conn.close()

        response_data = {
            'registered_students': registered_students,
            'current_sitins': current_sitins,
            'total_sitins': total_sitins,
            'purpose_stats': purpose_stats,
            'lab_stats': lab_stats
        }
        print("Sending response:", response_data)
        return jsonify(response_data)

    except Exception as e:
        print(f"Error in staff statistics: {str(e)}")
        return jsonify({'error': str(e)}), 500

@staff_bp.route('/feedbacks')
def get_feedbacks():
    if 'IDNO' not in session or session.get('USER_TYPE') != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT 
                f.FEEDBACK_ID,
                f.RECORD_ID,
                f.FEEDBACK_TEXT,
                f.DATE_SUBMITTED,
                u.IDNO as USER_IDNO,
                CONCAT(u.FIRSTNAME, ' ', u.LASTNAME) as STUDENT_NAME,
                l.LAB_NAME,
                s.DATE as SESSION_DATE,
                p.PURPOSE_NAME
            FROM FEEDBACKS f
            JOIN SIT_IN_RECORDS s ON f.RECORD_ID = s.RECORD_ID
            JOIN USERS u ON f.USER_IDNO = u.IDNO
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            ORDER BY f.DATE_SUBMITTED DESC
        """
        
        print("Executing feedbacks query:", query)
        cursor.execute(query)
        rows = cursor.fetchall()
        print(f"Found {len(rows)} feedbacks")
        
        feedbacks = []
        for row in rows:
            feedback_data = {
                'feedback_id': row[0],
                'record_id': row[1],
                'feedback_text': row[2],
                'date_submitted': row[3].strftime("%Y-%m-%d %H:%M:%S") if row[3] else None,
                'user_idno': row[4],
                'student_name': row[5],
                'lab_name': row[6],
                'session_date': row[7].strftime("%Y-%m-%d %H:%M:%S") if row[7] else None,
                'purpose_name': row[8],
                'PROFILE_PICTURE_URL': url_for('user.get_profile_picture', idno=row[4], _external=True)
            }
            print(f"Processing feedback for student {row[4]}: {feedback_data}")
            feedbacks.append(feedback_data)
        
        print("Returning feedbacks data:", feedbacks)
        return jsonify(feedbacks)
    except Exception as e:
        print(f"Error in get_feedbacks: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/generate_reports', methods=['GET'])
def generate_reports():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return redirect(url_for('auth.login'))

    # Get parameters from request
    report_type = request.args.get('type')
    report_format = request.args.get('format', 'pdf').lower()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    lab_id = request.args.get('lab_id')
    purpose_id = request.args.get('purpose_id')
    status = request.args.get('status')  # Add status filter
    session_status = request.args.get('session')  # Add session status filter
    
    # Validate required parameters
    if not report_type:
        return jsonify({'error': 'Report type is required'}), 400

    # Get lab name if lab_id is provided
    lab_name = None
    if lab_id:
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT LAB_NAME FROM LABORATORIES WHERE LAB_ID = %s", (lab_id,))
            result = cursor.fetchone()
            if result:
                lab_name = result[0]
        finally:
            cursor.close()
            conn.close()

    filters = {
        'Start Date': start_date,
        'End Date': end_date,
        'Laboratory': lab_name if lab_name else 'All Labs',
        'Status': status if status else 'All Statuses',
        'Session Status': session_status if session_status else 'All Sessions'
    }

    # Initialize report generator
    report_gen = ReportGenerator()

    try:
        # Prepare data based on report type
        if report_type == 'statistics':
            title = "Laboratory Usage Statistics Report"
            data = get_statistics_data(start_date, end_date, lab_id)
        elif report_type == 'feedback':
            title = "Student Feedback Report"
            data = get_feedback_data(start_date, end_date, lab_id)
        elif report_type in ['sit_in', 'sit_ins']:
            title = "Sit-in Records Report"
            data = get_sit_in_data(start_date, end_date, lab_id, purpose_id, status, session_status)
        else:
            return jsonify({'error': f'Invalid report type: {report_type}'}), 400

        # Generate report in requested format
        if report_format == 'pdf':
            buffer = report_gen.generate_pdf(title, data, filters)
            mimetype = 'application/pdf'
            file_ext = 'pdf'
        elif report_format == 'excel':
            buffer = report_gen.generate_excel(title, data, filters)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            file_ext = 'xlsx'
        elif report_format == 'csv':
            buffer = report_gen.generate_csv(title, data, filters)
            mimetype = 'text/csv'
            file_ext = 'csv'
        else:
            return jsonify({'error': f'Invalid format type: {report_format}'}), 400

        if not buffer:
            return jsonify({'error': 'Failed to generate report'}), 500

        # Send the file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_report_{timestamp}.{file_ext}"
        
        return send_file(
            buffer,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating report: {str(e)}")
        return jsonify({'error': str(e)}), 500

def get_statistics_data(start_date, end_date, lab_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Base query conditions
        conditions = []
        params = []
        
        if start_date:
            conditions.append("DATE >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("DATE <= %s")
            params.append(end_date)
        if lab_id:
            conditions.append("LAB_ID = %s")
            params.append(lab_id)
            
        where_clause = " AND ".join(conditions) if conditions else "1=1"

        # Get summary statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT USER_IDNO) as total_students,
                COUNT(*) as total_sessions
            FROM SIT_IN_RECORDS
            WHERE """ + where_clause, params)
        
        summary = cursor.fetchone()
        
        # Get lab usage statistics
        cursor.execute("""
            SELECT l.LAB_NAME, COUNT(*) as count
            FROM SIT_IN_RECORDS s
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            WHERE """ + where_clause + """
            GROUP BY l.LAB_NAME
            ORDER BY count DESC""", params)
        
        lab_stats = cursor.fetchall()
        
        # Get purpose statistics
        cursor.execute("""
            SELECT p.PURPOSE_NAME, COUNT(*) as count
            FROM SIT_IN_RECORDS s
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            WHERE """ + where_clause + """
            GROUP BY p.PURPOSE_NAME
            ORDER BY count DESC""", params)
        
        purpose_stats = cursor.fetchall()
        
        # Prepare data for report
        headers = ['Metric', 'Value']
        rows = [
            ['Total Students', str(summary[0])],
            ['Total Sessions', str(summary[1])],
            ['', ''],  # Empty row for spacing
            ['Laboratory Usage', '']
        ]
        
        # Add lab statistics
        for lab in lab_stats:
            rows.append([lab[0], str(lab[1])])
        
        rows.append(['', ''])  # Empty row for spacing
        rows.append(['Purpose Distribution', ''])
        
        # Add purpose statistics
        for purpose in purpose_stats:
            rows.append([purpose[0], str(purpose[1])])
            
        return [headers] + rows
        
    finally:
        cursor.close()
        conn.close()

def get_feedback_data(start_date, end_date, lab_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Base query conditions
        conditions = ["f.DATE_SUBMITTED IS NOT NULL"]
        params = []
        
        if start_date:
            conditions.append("f.DATE_SUBMITTED >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("f.DATE_SUBMITTED <= %s")
            params.append(end_date)
        if lab_id:
            conditions.append("s.LAB_ID = %s")
            params.append(lab_id)
            
        where_clause = " AND ".join(conditions)

        query = f"""
            SELECT 
                DATE(f.DATE_SUBMITTED) as date,
                l.LAB_NAME,
                u.IDNO,
                f.FEEDBACK_TEXT
            FROM FEEDBACKS f
            JOIN SIT_IN_RECORDS s ON f.RECORD_ID = s.RECORD_ID
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN USERS u ON f.USER_IDNO = u.IDNO
            WHERE {where_clause}
            ORDER BY f.DATE_SUBMITTED DESC
        """
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # Prepare data for report
        headers = ['Date', 'Laboratory', 'Student ID', 'Comments']
        rows = []
        
        for row in results:
            rows.append([
                row[0].strftime('%Y-%m-%d'),  # Date
                row[1],                        # Laboratory
                row[2],                        # Student ID
                row[3]                         # Comments
            ])
            
        return [headers] + rows
        
    finally:
        cursor.close()
        conn.close()

def get_sit_in_data(start_date, end_date, lab_id, purpose_id=None, status=None, session_status=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Base query conditions
        conditions = []
        params = []
        
        if start_date:
            conditions.append("s.DATE >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("s.DATE <= %s")
            params.append(end_date)
        if lab_id:
            conditions.append("s.LAB_ID = %s")
            params.append(lab_id)
        if purpose_id:
            conditions.append("s.PURPOSE_ID = %s")
            params.append(purpose_id)
        if status and status != "All Statuses":
            conditions.append("s.STATUS = %s")
            params.append(status)
        if session_status and session_status != "All Sessions":
            conditions.append("s.SESSION = %s")
            params.append(session_status)
            
        where_clause = " AND ".join(conditions) if conditions else "1=1"

        query = f"""
            SELECT 
                DATE(s.DATE) as date,
                l.LAB_NAME,
                u.IDNO,
                CONCAT(u.FIRSTNAME, ' ', u.LASTNAME) as student_name,
                s.DATE as time_in,
                s.END_TIME as time_out,
                TIMESTAMPDIFF(
                    MINUTE, 
                    s.DATE, 
                    CASE 
                        WHEN s.END_TIME IS NULL THEN NOW()
                        ELSE s.END_TIME
                    END
                ) as duration,
                p.PURPOSE_NAME,
                s.STATUS,
                s.SESSION
            FROM SIT_IN_RECORDS s
            JOIN LABORATORIES l ON s.LAB_ID = l.LAB_ID
            JOIN USERS u ON s.USER_IDNO = u.IDNO
            JOIN PURPOSES p ON s.PURPOSE_ID = p.PURPOSE_ID
            WHERE {where_clause}
            ORDER BY s.DATE DESC
        """
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # Prepare data for report
        headers = ['Date', 'Laboratory', 'Student ID', 'Student Name', 'Time In', 'Time Out', 'Duration (mins)', 'Purpose', 'Status', 'Session']
        rows = []
        
        for row in results:
            rows.append([
                row[0].strftime('%Y-%m-%d'),                          # Date
                row[1],                                               # Laboratory
                row[2],                                               # Student ID
                row[3],                                               # Student Name
                row[4].strftime('%Y-%m-%d %H:%M:%S'),                # Time In
                row[5].strftime('%Y-%m-%d %H:%M:%S') if row[5] else 'Ongoing',  # Time Out
                row[6],                                               # Duration
                row[7],                                               # Purpose
                row[8],                                               # Status
                row[9]                                                # Session
            ])
            
        return [headers] + rows
        
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/search_student/<string:student_id>')
def search_student(student_id):
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT IDNO, FIRSTNAME, LASTNAME, COURSE, YEAR, EMAIL 
            FROM USERS 
            WHERE IDNO = %s AND USER_TYPE = 'STUDENT'
        """, (student_id,))
        
        student = cursor.fetchone()
        if not student:
            return jsonify({'error': 'Student not found'}), 404

        return jsonify({
            'IDNO': student[0],
            'FIRSTNAME': student[1],
            'LASTNAME': student[2],
            'COURSE': student[3],
            'YEAR': student[4],
            'EMAIL': student[5]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/generate_report', methods=['GET'])
def generate_report():
    try:
        report_type = request.args.get('type', 'statistics')
        format_type = request.args.get('format', 'pdf')
        
        # Get current timestamp for filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if report_type == 'statistics':
            # Fetch statistics data
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT s.STUDENT_ID) as registered_students,
                    COUNT(CASE WHEN r.END_TIME IS NULL THEN 1 END) as active_sessions,
                    COUNT(*) as total_sessions,
                    l.LAB_NAME,
                    COUNT(r.RECORD_ID) as lab_sessions,
                    p.PURPOSE_NAME,
                    COUNT(r.RECORD_ID) as purpose_sessions
                FROM USERS s
                LEFT JOIN SIT_IN_RECORDS r ON s.IDNO = r.STUDENT_ID
                LEFT JOIN LABORATORIES l ON r.LAB_ID = l.LAB_ID
                LEFT JOIN PURPOSES p ON r.PURPOSE_ID = p.PURPOSE_ID
                WHERE s.USER_TYPE = 'Student'
                GROUP BY l.LAB_NAME, p.PURPOSE_NAME
            """)
            
            data = cursor.fetchall()
            
            if format_type == 'pdf':
                # Create PDF
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                elements = []
                
                # Title
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30
                )
                elements.append(Paragraph("System Statistics Report", title_style))
                elements.append(Spacer(1, 12))
                
                # Summary Table
                summary_data = [['Metric', 'Value']]
                summary_data.extend([
                    ['Registered Students', str(data[0][0])],
                    ['Active Sessions', str(data[0][1])],
                    ['Total Sessions', str(data[0][2])]
                ])
                
                # Create summary table
                summary_table = Table(summary_data, colWidths=[300, 200])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))
                
                # Lab Statistics
                elements.append(Paragraph("Laboratory Usage Statistics", styles['Heading2']))
                elements.append(Spacer(1, 12))
                
                lab_data = [['Laboratory', 'Number of Sessions']]
                lab_stats = {}
                for row in data:
                    if row[3]:  # LAB_NAME
                        lab_stats[row[3]] = row[4]  # lab_sessions
                
                lab_data.extend([[lab, str(sessions)] for lab, sessions in lab_stats.items()])
                
                lab_table = Table(lab_data, colWidths=[300, 200])
                lab_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(lab_table)
                elements.append(Spacer(1, 20))
                
                # Purpose Statistics
                elements.append(Paragraph("Purpose Statistics", styles['Heading2']))
                elements.append(Spacer(1, 12))
                
                purpose_data = [['Purpose', 'Number of Sessions']]
                purpose_stats = {}
                for row in data:
                    if row[5]:  # PURPOSE_NAME
                        purpose_stats[row[5]] = row[6]  # purpose_sessions
                
                purpose_data.extend([[purpose, str(sessions)] for purpose, sessions in purpose_stats.items()])
                
                purpose_table = Table(purpose_data, colWidths=[300, 200])
                purpose_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(purpose_table)
                
                # Build PDF
                doc.build(elements)
                buffer.seek(0)
                
                # Send PDF file
                return send_file(
                    buffer,
                    download_name=f'statistics_report_{timestamp}.pdf',
                    as_attachment=True,
                    mimetype='application/pdf'
                )
                
            elif format_type == 'csv':
                # Create CSV
                output = StringIO()
                writer = csv.writer(output)
                
                # Write summary
                writer.writerow(['System Statistics Report'])
                writer.writerow(['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([])
                writer.writerow(['Summary'])
                writer.writerow(['Registered Students', data[0][0]])
                writer.writerow(['Active Sessions', data[0][1]])
                writer.writerow(['Total Sessions', data[0][2]])
                writer.writerow([])
                
                # Write lab statistics
                writer.writerow(['Laboratory Statistics'])
                writer.writerow(['Laboratory', 'Number of Sessions'])
                lab_stats = {}
                for row in data:
                    if row[3]:
                        lab_stats[row[3]] = row[4]
                for lab, sessions in lab_stats.items():
                    writer.writerow([lab, sessions])
                writer.writerow([])
                
                # Write purpose statistics
                writer.writerow(['Purpose Statistics'])
                writer.writerow(['Purpose', 'Number of Sessions'])
                purpose_stats = {}
                for row in data:
                    if row[5]:
                        purpose_stats[row[5]] = row[6]
                for purpose, sessions in purpose_stats.items():
                    writer.writerow([purpose, sessions])
                
                output.seek(0)
                return send_file(
                    BytesIO(output.getvalue().encode('utf-8')),
                    download_name=f'statistics_report_{timestamp}.csv',
                    as_attachment=True,
                    mimetype='text/csv'
                )
                
            else:  # Excel format
                wb = Workbook()
                
                # Summary sheet
                ws = wb.active
                ws.title = "Summary"
                
                # Title
                ws['A1'] = "System Statistics Report"
                ws['A1'].font = Font(size=14, bold=True)
                ws.merge_cells('A1:B1')
                
                # Summary data
                ws['A3'] = "Metric"
                ws['B3'] = "Value"
                ws['A4'] = "Registered Students"
                ws['B4'] = data[0][0]
                ws['A5'] = "Active Sessions"
                ws['B5'] = data[0][1]
                ws['A6'] = "Total Sessions"
                ws['B6'] = data[0][2]
                
                # Style summary header
                for cell in ws['A3:B3'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Laboratory Statistics sheet
                ws_lab = wb.create_sheet("Laboratory Statistics")
                ws_lab['A1'] = "Laboratory"
                ws_lab['B1'] = "Number of Sessions"
                
                row = 2
                lab_stats = {}
                for d in data:
                    if d[3]:
                        lab_stats[d[3]] = d[4]
                for lab, sessions in lab_stats.items():
                    ws_lab[f'A{row}'] = lab
                    ws_lab[f'B{row}'] = sessions
                    row += 1
                
                # Style lab header
                for cell in ws_lab['A1:B1'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Purpose Statistics sheet
                ws_purpose = wb.create_sheet("Purpose Statistics")
                ws_purpose['A1'] = "Purpose"
                ws_purpose['B1'] = "Number of Sessions"
                
                row = 2
                purpose_stats = {}
                for d in data:
                    if d[5]:
                        purpose_stats[d[5]] = d[6]
                for purpose, sessions in purpose_stats.items():
                    ws_purpose[f'A{row}'] = purpose
                    ws_purpose[f'B{row}'] = sessions
                    row += 1
                
                # Style purpose header
                for cell in ws_purpose['A1:B1'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Auto-adjust column widths
                for worksheet in [ws, ws_lab, ws_purpose]:
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Save to buffer
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                return send_file(
                    excel_buffer,
                    download_name=f'statistics_report_{timestamp}.xlsx',
                    as_attachment=True,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        
        conn.close()
        return jsonify({'error': 'Invalid report type'})
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        return jsonify({'error': str(e)}), 500

@staff_bp.route('/all_students')
def get_all_students():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get all students with their basic information
        cursor.execute("""
            SELECT u.IDNO, u.FIRSTNAME, u.LASTNAME, u.COURSE, u.YEAR, u.EMAIL,
                   COALESCE(s.SIT_IN_COUNT, 0) as REMAINING_SESSIONS
            FROM USERS u
            LEFT JOIN SIT_IN_LIMITS s ON u.IDNO = s.USER_IDNO
            WHERE u.USER_TYPE = 'STUDENT'
            ORDER BY u.LASTNAME, u.FIRSTNAME
        """)
        
        students = cursor.fetchall()
        result = []
        for student in students:
            # Get profile picture URL
            profile_url = url_for('user.get_profile_picture', idno=student[0], _external=True)
            
            result.append({
                'IDNO': student[0],
                'FIRSTNAME': student[1],
                'LASTNAME': student[2],
                'COURSE': student[3],
                'YEAR': student[4],
                'EMAIL': student[5],
                'REMAINING_SESSIONS': student[6],
                'PROFILE_PICTURE_URL': profile_url
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/reset_student_session/<string:student_id>', methods=['POST'])
def reset_student_session(student_id):
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    if not student_id or student_id == 'undefined':
        return jsonify({'error': 'Invalid student ID'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get student's course
        cursor.execute("""
            SELECT IDNO, COURSE 
            FROM USERS 
            WHERE IDNO = %s AND USER_TYPE = 'STUDENT'
        """, (student_id,))
        result = cursor.fetchone()
        
        if not result:
            return jsonify({'error': 'Student not found'}), 404

        student_id = result[0]  # Use the verified student ID
        course = result[1]
        
        # Set session limit based on course
        session_limit = 30 if course in ['BSIT', 'BSCS'] else 15

        # First check if record exists
        cursor.execute("SELECT USER_IDNO FROM SIT_IN_LIMITS WHERE USER_IDNO = %s", (student_id,))
        exists = cursor.fetchone()

        if exists:
            # Update existing record
            cursor.execute("""
                UPDATE SIT_IN_LIMITS 
                SET SIT_IN_COUNT = %s 
                WHERE USER_IDNO = %s
            """, (session_limit, student_id))
        else:
            # Insert new record
            cursor.execute("""
                INSERT INTO SIT_IN_LIMITS (USER_IDNO, SIT_IN_COUNT)
                VALUES (%s, %s)
            """, (student_id, session_limit))

        conn.commit()
        return jsonify({
            'success': True, 
            'message': f'Sessions reset to {session_limit}',
            'remaining_sessions': session_limit
        })
    except Exception as e:
        conn.rollback()
        print(f"Error in reset_student_session: {str(e)}")  # Add logging
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/reset_all_sessions', methods=['POST'])
def reset_all_sessions():
    if 'IDNO' not in session or session['USER_TYPE'] != 'STAFF':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # First, get all students and their courses
        cursor.execute("""
            SELECT IDNO, COURSE 
            FROM USERS 
            WHERE USER_TYPE = 'STUDENT'
        """)
        students = cursor.fetchall()
        
        if not students:
            return jsonify({'error': 'No students found'}), 404

        success_count = 0
        error_count = 0

        # Process each student
        for student_id, course in students:
            try:
                session_limit = 30 if course in ['BSIT', 'BSCS'] else 15
                
                # Check if record exists
                cursor.execute("SELECT USER_IDNO FROM SIT_IN_LIMITS WHERE USER_IDNO = %s", (student_id,))
                exists = cursor.fetchone()
                
                if exists:
                    # Update existing record
                    cursor.execute("""
                        UPDATE SIT_IN_LIMITS 
                        SET SIT_IN_COUNT = %s 
                        WHERE USER_IDNO = %s
                    """, (session_limit, student_id))
                else:
                    # Insert new record
                    cursor.execute("""
                        INSERT INTO SIT_IN_LIMITS (USER_IDNO, SIT_IN_COUNT)
                        VALUES (%s, %s)
                    """, (student_id, session_limit))
                
                success_count += 1
                
            except Exception as e:
                print(f"Error processing student {student_id}: {str(e)}")  # Add logging
                error_count += 1
                continue  # Continue with next student even if one fails

        conn.commit()
        return jsonify({
            'success': True,
            'message': f'Reset completed. Success: {success_count}, Errors: {error_count}',
            'details': {
                'success_count': success_count,
                'error_count': error_count,
                'total_processed': len(students)
            }
        })
    except Exception as e:
        conn.rollback()
        print(f"Error in reset_all_sessions: {str(e)}")  # Add logging
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@staff_bp.route('/top_students')
def get_top_students():
    print("\n=== Starting get_top_students endpoint ===")
    if 'IDNO' not in session or session['USER_TYPE'] not in ['STAFF', 'ADMIN']:
        print("Unauthorized access attempt")
        return jsonify({'error': 'Unauthorized'}), 401

    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        print("Executing SQL query for top students...")
        cursor.execute("""
            SELECT 
                u.IDNO,
                CONCAT(u.FIRSTNAME, ' ', u.LASTNAME) as STUDENT_NAME,
                u.COURSE,
                COUNT(s.RECORD_ID) as TOTAL_SITINS
            FROM USERS u
            JOIN SIT_IN_RECORDS s ON u.IDNO = s.USER_IDNO
            WHERE u.USER_TYPE = 'STUDENT'
            GROUP BY u.IDNO, u.FIRSTNAME, u.LASTNAME, u.COURSE
            ORDER BY TOTAL_SITINS DESC
            LIMIT 5
        """)
        
        print("Fetching results...")
        results = cursor.fetchall()
        print(f"Found {len(results)} students")
        
        top_students = []
        for row in results:
            print(f"Processing student: {row[0]} - {row[1]}")
            # Get profile picture for each student
            cursor.execute("SELECT PROFILE_PICTURE FROM USERS WHERE IDNO = %s", (row[0],))
            profile_result = cursor.fetchone()
            profile_pic = f"/user/get_profile_picture/{row[0]}"
            
            student_data = {
                'student_id': row[0],
                'student_name': row[1],
                'course': row[2],
                'total_sitins': row[3],
                'profile_picture_url': profile_pic
            }
            print(f"Student data: {student_data}")
            top_students.append(student_data)
        
        print(f"Returning {len(top_students)} students")
        print("=== Finished get_top_students endpoint ===\n")
        return jsonify(top_students)
    except Exception as e:
        print(f"Error in get_top_students: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close() 